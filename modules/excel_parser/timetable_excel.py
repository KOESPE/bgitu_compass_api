import datetime

import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from config import EXCEL_DIRECTORY
from data import WEEKDAY_INDEX
from database.base import manage_subjects, manage_groups, db_init, insert_schedule, db_reset_schedules, db_drop_tables
from modules.excel_parser.excel_parser_functions import make_dict_day, extract_numbers, standardize_names
from modules.schedule_managment.weekly_updates import weekly_management_schedule


async def initialize_excel_schedules(full_reset=False,
                                     reset_schedules=False,
                                     reset_schedules_and_subjects=False):
    """
    Инициализация расписания из excel файлов в db.groups.rawSchedule

    :param full_reset: drop tables + init tables
    :param reset_schedules: Delete Lessons
    :return:
    """
    if full_reset:
        await db_drop_tables()
        await db_init()

    if reset_schedules:
        await db_reset_schedules()

    if reset_schedules_and_subjects:
        await db_reset_schedules(and_subjects=True)

    excel_files = [file for file in EXCEL_DIRECTORY.iterdir() if '.xlsx' in file.name]

    for file in excel_files:
        wb = openpyxl.load_workbook(file)
        sheet = wb.worksheets[0]

        await process_sheet(sheet)

    # В контексте нового учебного года это может создавать ошибки
    if full_reset or reset_schedules or reset_schedules_and_subjects:
        await weekly_management_schedule()


async def process_sheet(sheet):
    """Обрабатывает лист Excel и извлекает расписание для каждой группы."""

    for group_column in range(4, sheet.max_column + 1):
        group_name = parse_group_name(sheet, group_column)
        if group_name is None:  # Пустая ячейка
            continue

        group_id = await manage_groups(group_name)

        schedule_week = await parse_group_schedule(sheet, group_id, group_column)
        await insert_schedule(group_id, schedule_week)


def parse_group_name(sheet, group_column):
    """Извлекает имя группы из листа Excel."""

    group_name = parse_cell(sheet=sheet, row=3, col=group_column)
    if group_name is None:
        return None

    group_name = group_name.upper()
    group_name = group_name.replace(' ', '')
    group_name = group_name.replace('/', '-')

    if group_name.lower() in ('а', 'б'):  # Подгруппы
        group_name = (parse_cell(sheet, row=2, col=group_column).split())[0] + f'({group_name})'

    if 'спо' in group_name.lower():
        group_name = group_name.replace('спо', 'СПО')

    return group_name


async def parse_group_schedule(sheet, group_id, group_column):
    """Парсит расписание для одной группы."""

    schedule_week = {"first_week": {}, "second_week": {}}
    schedule_per_day_first_week = []
    schedule_per_day_second_week = []

    weekday_last = parse_cell(sheet, row=4, col=1)
    lesson_step = 2

    for group_row in range(4, sheet.max_row + 1, lesson_step):
        weekday_now = parse_cell(sheet, row=group_row, col=1)

        if weekday_now != weekday_last:
            schedule_week['second_week'][WEEKDAY_INDEX[weekday_last.lower()]] = schedule_per_day_first_week
            schedule_week['first_week'][WEEKDAY_INDEX[weekday_last.lower()]] = schedule_per_day_second_week

            if weekday_now:
                weekday_last = weekday_now
                schedule_per_day_first_week = []
                schedule_per_day_second_week = []

        lesson_first_week = parse_day(sheet, row=group_row, col=group_column)
        lesson_second_week = parse_day(sheet, row=group_row + 1, col=group_column)

        if lesson_first_week:
            subject_id = await manage_subjects(lesson_first_week.get('subjectName'), group_id)
            day_template = make_dict_day(data=lesson_first_week, subject_id=subject_id)
            schedule_per_day_first_week.append(day_template)

        if lesson_second_week:
            subject_id = await manage_subjects(lesson_second_week.get('subjectName'), group_id)
            day_template = make_dict_day(data=lesson_second_week, subject_id=subject_id)
            schedule_per_day_second_week.append(day_template)

    schedule_week['second_week'][WEEKDAY_INDEX[weekday_last.lower()]] = schedule_per_day_first_week
    schedule_week['first_week'][WEEKDAY_INDEX[weekday_last.lower()]] = schedule_per_day_second_week
    return schedule_week


def parse_cell(sheet: Worksheet, row, col, using_merged=True):
    cell = sheet.cell(row=row, column=col)
    if using_merged:
        if isinstance(cell, MergedCell):
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
        else:
            cell = sheet.cell(row, col)
        value = cell.value  # Plain text
    else:
        value = cell.value
    return value


col_weekday = 1
col_time_1_building = 2
col_time_2_building = 3


def parse_day(sheet: Worksheet, row, col):
    """
    Парсинг в пределах одной пары
    :param sheet: Таблица из файла
    :param row: Строка
    :param col: Столбец
    """
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    else:
        cell = sheet.cell(row, col)
    value = cell.value  # Текст ячейки

    if (value is None) or (value == '') or (value == ' '):  # Обработка пустоты
        return None

    # Проверка на второй корпус по цвету текста
    building = 1
    if cell.font.color is not None:
        if cell.font.color.rgb == 'FFFF0000' or cell.font.color.indexed == 10:
            building = 2

    # Проверка на тип занятия по bold
    if cell.font.bold:
        is_lecture = True
    else:
        is_lecture = False

    # Следуя из корпуса получаем время
    if building == 2:
        lesson_time = parse_cell(sheet, row, col_time_2_building)
    else:
        lesson_time = parse_cell(sheet, row, col_time_1_building)

    try:
        lesson_time = lesson_time.split('-')
        start_at = datetime.datetime.strptime(lesson_time[0], '%H.%M').time()
        end_at = datetime.datetime.strptime(lesson_time[1], '%H.%M').time()
    except (ValueError, AttributeError) as e:  # Какая-то некорректная ячейка, не смог отловить в свое время ошибку
        return None

    value: str = value.replace("\n", " ")  # Ненужные переносы

    # Обработка фраз "консп." и временное удаления для корректной работы алгоритмы
    strange_phrase = ''
    if '(' in value and ')' in value:
        strange_phrase = value[value.find('('):value.find(')') + 1]
        value = value.replace(strange_phrase, '')

    # В строчках бывает очень много пробелов
    for i in range(2, 30):
        value = value.replace(' ' * i, ' ')

    # Основа алгоритма для вытаскивания данных из ячейки
    str_list = value.split(' ')
    str_list = list(filter(None, str_list))  # Убираем '' из списка
    str_list_classrooms = str_list.copy()  # Сохранение оригинала для логики преподавателей
    if len(str_list) == 0:  # В некоторых ячейках есть непонятный невидимый символ переноса
        return None

    classrooms_algorithm = extract_numbers(str_list)
    classrooms = classrooms_algorithm['classrooms']
    str_list = classrooms_algorithm['str_list_no_classrooms']

    if bool(classrooms):  # Самая обычная ячейка
        if len(classrooms) == 1:
            classroom = classrooms[0]

            """ Обработка спец символов:
            / — разделение на чет/нечет - определяем по клетке
            : или ; - просто второй кабинет, ничего не делаем """
            if '/' in classroom:
                lesson_time = parse_cell(sheet, row, col_time_1_building)
                next_lesson_time = parse_cell(sheet, row + 1, col_time_1_building)
                if lesson_time == next_lesson_time:  # Значит это первая неделя, возвращаем число перед тире
                    classroom = classroom.split('/')[0] + f'({classroom.split("/")[1]})'
                else:
                    classroom = classroom.split('/')[1] + f'({classroom.split("/")[0]})'

            # Странное написание преподавателя — Ситуация вида "СальниковаН.А."
            if str_list_classrooms[-2].replace('/', '').replace(':', '').replace(';', '').isdigit():
                teacher = str_list[-1]
                del str_list[-1:]
            else:
                teacher = str_list[-2] + ' ' + str_list[-1]
                del str_list[-2:]

        else:  # Иногда бывает два кабинета в одной ячейке. !!! В этом случае первое значение — группа А, второе — Б
            right_cell_value = parse_cell(sheet, row, col + 1)
            if cell.value == right_cell_value:
                classroom = classrooms[0]

                str_list.remove(classrooms[0])
                str_list.remove(classrooms[1])

                teacher = str_list[-4] + ' ' + str_list[-3]
            else:
                classroom = classrooms[1]

                str_list.remove(classrooms[0])
                str_list.remove(classrooms[1])

                teacher = str_list[-2] + ' ' + str_list[-1]
            del str_list[-4:]
        teacher = standardize_names(teacher)  # Правим написание преподавателя

        subject_name = ' '.join(str_list)
        if classroom.upper() == 'ДОТ':
            building = 'ДОТ'
            classroom = 'ДОТ'

    else:  # Преподаватель отсутствует и строчка из себя представляет название предмета
        subject_name = ' '.join(str_list)
        classroom = ''
        teacher = ''
        is_lecture = False

    if len(strange_phrase) > 0:
        subject_name += ' ' + strange_phrase
    subject_name = ' '.join(str_list)

    return {
        "subjectName": subject_name,
        "building": str(building),
        "startAt": start_at,
        "endAt": end_at,
        "classroom": classroom,
        "teacher": teacher,
        "isLecture": is_lecture
    }
